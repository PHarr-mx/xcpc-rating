from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from importer.awards import assign_award, resolve_award_thresholds
from importer.models import AwardThresholds, XcpcioParsedContest, XcpcioStandingRow

_MEDAL_MAP = {
    "gold": "gold",
    "silver": "silver",
    "bronze": "bronze",
    "honorable": "honorable",
}


def _normalize_header(value: object) -> str:
    return str(value or "").strip().lower()


def _header_index(headers: tuple[object, ...]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for index, header in enumerate(headers):
        key = _normalize_header(header)
        if key:
            mapping[key] = index
    return mapping


def _cell(row: tuple[object, ...], headers: dict[str, int], name: str) -> object | None:
    index = headers.get(name.lower())
    if index is None or index >= len(row):
        return None
    return row[index]


def _parse_int(value: object) -> int:
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    return int(str(value).strip())


def _parse_bool_flag(value: object) -> bool:
    return str(value or "").strip().upper() == "Y"


def _parse_members(row: tuple[object, ...], headers: dict[str, int]) -> list[str]:
    members: list[str] = []
    for key in ("member1", "member2", "member3"):
        raw = _cell(row, headers, key)
        if raw is None:
            continue
        name = str(raw).strip()
        if name:
            members.append(name)
    return members


def _parse_award(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text == "-":
        return None
    return _MEDAL_MAP.get(text.lower(), text.lower())


def _count_total_teams(sheet) -> int:
    rows = list(sheet.iter_rows(min_row=3, values_only=True))
    return sum(1 for row in rows if row and row[0] not in (None, "", "Rank"))


def _count_problem_columns(headers: tuple[object, ...]) -> int:
    """统计表头中连续的单字母题号列（A、B、C…），遇非题号列即停止。"""
    count = 0
    for header in headers:
        text = str(header or "").strip()
        if len(text) == 1 and text.isalpha() and text.upper() == text:
            count += 1
        elif count > 0:
            break
    return count


def _read_problem_count(sheet) -> int:
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        return 0
    return _count_problem_columns(rows[1])


def _parse_sheet_rows(
    sheet,
    *,
    school_organizations: set[str] | None = None,
    include_unofficial: bool = True,
    read_medal: bool = False,
) -> list[XcpcioStandingRow]:
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    headers = _header_index(rows[1])
    results: list[XcpcioStandingRow] = []

    for row in rows[2:]:
        if not row or row[0] in (None, ""):
            continue

        organization = str(_cell(row, headers, "organization") or "").strip()
        if school_organizations is not None and organization not in school_organizations:
            continue

        unofficial = _parse_bool_flag(_cell(row, headers, "unofficial"))
        if unofficial and not include_unofficial:
            continue

        members = _parse_members(row, headers)
        if not members:
            continue

        school_rank_raw = _cell(row, headers, "organization rank")
        school_rank = _parse_int(school_rank_raw) if school_rank_raw not in (None, "") else None
        if school_rank == 0:
            school_rank = None

        award = _parse_award(_cell(row, headers, "medal")) if read_medal else None

        results.append(
            XcpcioStandingRow(
                rank=_parse_int(_cell(row, headers, "rank")),
                school_rank=school_rank,
                organization=organization,
                team_name=str(_cell(row, headers, "team") or "").strip(),
                solved=_parse_int(_cell(row, headers, "solved")),
                penalty=_parse_int(_cell(row, headers, "penalty")),
                award=award,
                members=members,
                unofficial=unofficial,
            )
        )
    return results


def _to_standing_scores(rows: list[XcpcioStandingRow]):
    from importer.awards import StandingScore

    return [
        StandingScore(rank=row.rank, solved=row.solved, penalty=row.penalty, award=row.award)
        for row in rows
    ]


def _apply_awards(
    school_rows: list[XcpcioStandingRow],
    thresholds: AwardThresholds,
) -> list[XcpcioStandingRow]:
    awarded: list[XcpcioStandingRow] = []
    for row in school_rows:
        award = assign_award(row.solved, row.penalty, thresholds)
        if award is None:
            continue
        awarded.append(row.model_copy(update={"award": award}))
    return sorted(awarded, key=lambda item: item.rank)


def parse_xcpcio_xlsx(
    path: Path | str,
    *,
    school_organizations: list[str],
    standings_sheet: str = "正式组",
    total_teams_sheet: str = "所有队伍",
    include_unofficial: bool = True,
) -> XcpcioParsedContest:
    path = Path(path)
    school_set = set(school_organizations)
    workbook = load_workbook(path, read_only=True, data_only=True)

    try:
        if total_teams_sheet not in workbook.sheetnames:
            raise ValueError(f"缺少工作表: {total_teams_sheet}")
        if standings_sheet not in workbook.sheetnames:
            raise ValueError(f"缺少工作表: {standings_sheet}")

        total_sheet = workbook[total_teams_sheet]
        formal_sheet = workbook[standings_sheet]
        title = str(total_sheet["A1"].value or path.stem).strip()
        total_teams = _count_total_teams(total_sheet)
        total_problems = _read_problem_count(total_sheet)
        if total_problems <= 0:
            total_problems = _read_problem_count(formal_sheet)

        formal_rows = _parse_sheet_rows(
            formal_sheet,
            include_unofficial=True,
            read_medal=True,
        )
        all_rows = _parse_sheet_rows(
            total_sheet,
            include_unofficial=True,
        )
        school_rows = _parse_sheet_rows(
            total_sheet,
            school_organizations=school_set,
            include_unofficial=include_unofficial,
        )

        thresholds = resolve_award_thresholds(
            _to_standing_scores(formal_rows),
            _to_standing_scores(all_rows),
            total_teams=total_teams,
        )
        standings = _apply_awards(school_rows, thresholds)
    finally:
        workbook.close()

    if total_teams <= 0:
        raise ValueError("无法解析 total_teams")
    if total_problems <= 0:
        raise ValueError("无法解析 total_problems")
    if not school_rows:
        raise ValueError("未匹配到任何本校队伍，请检查 school_organizations 或工作表名称")
    if not standings:
        raise ValueError("本校队伍均无金/银/铜奖，未写入任何成绩")

    return XcpcioParsedContest(
        title=title,
        total_teams=total_teams,
        total_problems=total_problems,
        standings_sheet=standings_sheet,
        total_teams_sheet=total_teams_sheet,
        standings=standings,
        school_teams_total=len(school_rows),
        award_thresholds=thresholds,
    )
